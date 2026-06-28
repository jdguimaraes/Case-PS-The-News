import nbformat as nbf

nb = nbf.v4.new_notebook()
cells = []
def md(text): cells.append(nbf.v4.new_markdown_cell(text))
def code(text): cells.append(nbf.v4.new_code_cell(text))

md("""# Palavritas — O que determina o retorno do jogador?
### Analise de retencao (D+1 e D30) — Analista de Dados, Produto & Growth | the news

Este notebook reune **codigo, raciocinio e resultado** lado a lado para cada etapa da analise,
do jeito que o case pediu. O documento `Analise_Palavritas_the_news.docx` traz a mesma analise
em linguagem executiva, sem codigo, para o Head de Produto. Aqui esta o "como" por tras de cada numero.

**Pergunta do case:** O que esta determinando se um usuario volta a jogar — e o que podemos
fazer para aumentar isso?

**Resposta curta, ja adiantada:** o retorno no dia seguinte (D+1) **nao tem driver relevante**
nos dados disponiveis — fica em ~22% em quase todo corte testado. Ja a retencao em 30 dias (D30)
tem 3 drivers claros e estatisticamente significativos: jogar de manha, abrir a newsletter antes
do jogo, e ter voltado no D+1. Combinados, newsletter + D+1 levam o D30 de 22% para 92%.
""")

# ---------- ENTREGA 1 ----------
md("""## Entrega 1 — Limpeza e diagnostico

Antes de qualquer analise, carregamos as 3 abas originais e investigamos problemas de qualidade
(`01_explore.py`, `02_quality.py`, `03_quality2.py`). Abaixo o codigo de carga e os principais
achados de qualidade, seguidos da limpeza efetivamente aplicada (`04_clean_and_analyze.py`).
""")

code("""import pandas as pd, numpy as np
path = r"../Case Analista de Dados — the news [Dataset Palavritas].xlsx"
xls = pd.read_excel(path, sheet_name=None)
for name, df in xls.items():
    print(name, df.shape)
""")

md("""### Problemas encontrados (resumo — detalhe linha a linha em `cleaning_log.txt`)

| # | Problema | Onde | Decisao |
|---|---|---|---|
| 1 | 1.198 linhas 100% duplicadas | sessions | drop_duplicates() |
| 2 | Device com casing inconsistente (`android`, `Android`, `IOS`...) | sessions | normalizado para `Android`/`iOS` |
| 3 | `word_date` em dois formatos (`YYYY-MM-DD` e `DD/MM/YYYY`) | sessions | parse com deteccao de formato por regex |
| 4 | `time_to_complete_sec` <= 0 (40 linhas) | sessions | anulado (NaN), linha mantida |
| 5 | `attempts` fora do range 1-6 ou `result` nulo (110 linhas, 0.28%) | sessions | descartado — corrupcao, nao recuperavel |
| 6 | `sessions.attempts` != contagem real em `attempts` (62 linhas, 0.16%) | sessions x attempts | so documentado, nao corrigido (efeito marginal) |
| 7 | 752 linhas duplicadas | attempts | drop_duplicates() |
| 8 | `guess` com tamanho != 5 (80 linhas) | attempts | descartado — corrupcao |
| 9 | Mojibake em `sector`/`salary_range`/`job_role`/`state` (encoding corrompido na origem) | user_profile | reconstruido manualmente por contexto (poucas palavras, todas reconhecveis) |
| 10 | 400 dos 1.200 usuarios de `sessions` sem registro em `user_profile` | sessions x user_profile | documentado como limitacao — qualquer corte por perfil cobre so 800/1200 usuarios (67%) |

**O que decidimos NAO tratar:** nulos legitimos de nao-resposta em `age_range` (117), `city` (297),
`salary_range` (193) — sao nao-resposta de pesquisa, nao erro de dado, e foram mantidos como NaN
para nao inflar nenhuma categoria artificialmente.
""")

code("""# Exemplo do que checamos: deteccao de formato de data misto
import re
def fmt(x):
    if re.match(r'^\\d{4}-\\d{2}-\\d{2}$', x): return 'YYYY-MM-DD'
    if re.match(r'^\\d{2}/\\d{2}/\\d{4}$', x): return 'DD/MM/YYYY'
    return 'other:'+x
s_raw = xls['palavritas_sessions']
print(s_raw['word_date'].apply(fmt).value_counts())
""")

code("""# Exemplo do mojibake na origem (nao e bug de leitura do pandas - confirmado lendo bytes crus)
from openpyxl import load_workbook
wb = load_workbook(path, read_only=True)
ws = wb['user_profile']
for r in list(ws.iter_rows(min_row=1, max_row=4, values_only=True)):
    print(r)
""")

md("""### Limpeza aplicada

Carregamos os CSVs ja limpos (saida de `04_clean_and_analyze.py`) para o resto da analise.
""")

code("""s = pd.read_csv('sessions_clean.csv', parse_dates=['word_date_parsed'])
u = pd.read_csv('user_profile_clean.csv')
a = pd.read_csv('attempts_clean.csv')
s['played_next_day'] = s['played_next_day'].astype(bool)
s['active_d30'] = s['active_d30'].astype(bool)
s['newsletter_open_before_game'] = s['newsletter_open_before_game'].astype(bool)
print(f"sessions: {len(s)} linhas | usuarios unicos: {s['user_id'].nunique()} | perfil disponivel: {u['user_id'].nunique()}")
""")

# ---------- ENTREGA 2 ----------
md("""## Entrega 2 — Analise

Testamos cada variavel candidata contra as duas metricas de retorno, usando:
- **Qui-quadrado** para variaveis categoricas (ex.: device, setor, horario em faixas)
- **Teste-t de Welch** (`equal_var=False`) para variaveis continuas (ex.: `session_hour`, `streak_day`)

Bar de decisao: **p < 0.05** (95% de confianca) para considerar um efeito real.
""")

code("""from scipy import stats
def chi2_test(df, col, target):
    ct = pd.crosstab(df[col], df[target])
    chi2, p, dof, _ = stats.chi2_contingency(ct)
    rate = df.groupby(col)[target].mean()
    n = df.groupby(col)[target].count()
    print(f"-- {col} vs {target} -- chi2={chi2:.2f}, p={p:.6f}, dof={dof}")
    print(pd.DataFrame({'rate':rate, 'n':n}).sort_values('rate', ascending=False))
    return p

def ttest_cont(df, col, target):
    g1 = df[df[target]==True][col].dropna()
    g0 = df[df[target]==False][col].dropna()
    t, p = stats.ttest_ind(g1, g0, equal_var=False)
    print(f"-- {col} vs {target} -- mean(True)={g1.mean():.2f} mean(False)={g0.mean():.2f} t={t:.2f} p={p:.6f}")
    return p

s['hour_bucket'] = pd.cut(s['session_hour'], bins=[5,9,12,17,21,24],
    labels=['manha cedo(6-9)','manha(10-12)','tarde(13-17)','noite(18-21)','madrugada(22-23)'])
""")

md("""### D+1 (retorno no dia seguinte) — testando horario, palavra, device, resultado, perfil...""")

code("""ttest_cont(s, 'session_hour', 'played_next_day')
chi2_test(s, 'hour_bucket', 'played_next_day')""")

code("""chi2_test(s, 'device', 'played_next_day')""")

code("""chi2_test(s, 'result', 'played_next_day')""")

code("""su = s.merge(u, on='user_id', how='left')
chi2_test(su.dropna(subset=['sector']), 'sector', 'played_next_day')""")

code("""chi2_test(su.dropna(subset=['salary_range']), 'salary_range', 'played_next_day')""")

code("""chi2_test(su.dropna(subset=['orders_food_delivery']), 'orders_food_delivery', 'played_next_day')""")

code("""chi2_test(s, 'newsletter_open_before_game', 'played_next_day')""")

md("""**Conclusao parcial:** nenhuma das variaveis acima passa de p=0.05 com efeito pratico relevante
para D+1 — a taxa fica colada em ~22% em praticamente todo corte. Ver `stats_results.txt` para a
lista completa (streak_day, time_to_complete_sec, newsletter_subscriber, plays_other_word_games
tambem testados, mesmo resultado).
""")

md("""### Aprofundando no D+1: a tabela `attempts` (146 mil linhas) explica algo que o resultado binario nao explica?

O dataset traz o detalhe de cada palpite (numero da tentativa, letras certas, posicoes certas).
Testamos duas hipoteses adicionais antes de aceitar que D+1 realmente nao tem driver:
1. **Velocidade de vitoria** — ganhar na 1a tentativa entusiasma mais que ganhar na 6a?
2. **Quase-acerto na derrota** — perder por pouco (4 de 5 letras certas) frustra/engaja mais que perder de longe?
3. **Efeitos compostos** — combinar variaveis (ganhar + jogar de manha; newsletter + ganhar) revela algo que cada uma isolada nao revela?
""")

code("""last_attempt = a.groupby('session_id')['attempt_number'].max().rename('attempts_used')
s2 = s.merge(last_attempt, on='session_id', how='left')
won = s2[s2['result'].astype(str).str.lower().str.contains('win|venceu|ganh', regex=True, na=False)]
print(f"sessoes vitoria: {len(won)} de {len(s2)}")
ttest_cont(won, 'attempts_used', 'played_next_day')
chi2_test(won, 'attempts_used', 'played_next_day')""")

code("""lost = s2[~s2.index.isin(won.index)]
last_try = a.sort_values('attempt_number').groupby('session_id').tail(1)[['session_id','correct_positions','correct_letters']]
lost2 = lost.merge(last_try, on='session_id', how='left')
lost2['near_miss'] = lost2['correct_positions'] >= 4
ttest_cont(lost2.dropna(subset=['correct_positions']), 'correct_positions', 'played_next_day')
chi2_test(lost2.dropna(subset=['near_miss']), 'near_miss', 'played_next_day')""")

code("""s2['won'] = won.index.to_series().reindex(s2.index).notna() if False else s2.index.isin(won.index)
s2['morning'] = s2['hour_bucket'].isin(['manha cedo(6-9)','manha(10-12)'])
print(s2.groupby(['won','morning'])['played_next_day'].agg(['mean','count']))
print()
print(s2.groupby(['newsletter_open_before_game','won'])['played_next_day'].agg(['mean','count']))""")

md("""**Conclusao:** nenhuma das tres hipoteses adicionais (velocidade, quase-acerto, combinacoes)
muda a taxa de D+1 de forma relevante — todas ficam entre 21% e 23%, sem significancia estatistica
(p > 0.08 em todos os casos, ver `attempts_analysis_results.txt`). Isso reforça, com mais rigor, a
conclusao original: **dentro deste dataset, o retorno no dia seguinte parece ser dirigido por algo
fora dos dados que temos** (ex.: notificacao push, habito pre-existente do usuario) — nao pela
experiencia de jogo em si.
""")

md("""### D30 (ativo em 30 dias) — aqui sim ha drivers claros""")

code("""chi2_test(s, 'hour_bucket', 'active_d30')""")

code("""chi2_test(s, 'newsletter_open_before_game', 'active_d30')""")

code("""chi2_test(s, 'played_next_day', 'active_d30')""")

code("""# O achado principal: efeito composto newsletter + D+1
combo = s.groupby(['newsletter_open_before_game','played_next_day'])['active_d30'].mean()
print(combo)
print()
print("D30 quando abre newsletter E volta no D+1:", f"{combo[(True,True)]*100:.1f}%")
print("D30 quando nao faz nenhuma das duas:      ", f"{combo[(False,False)]*100:.1f}%")""")

md("""**Causalidade vs correlacao:** `newsletter_open_before_game` provavelmente marca um usuario
*ja engajado*, nao causa o engajamento por si so. Tratamos isso como hipotese a testar (A/B test),
nao como fato estabelecido — ver Entrega 3.
""")

# ---------- ENTREGA 3 ----------
md("""## Entrega 3 — Proposta

**Hipotese:** Acredito que o habito de abrir a newsletter antes de jogar — combinado com voltar no
dia seguinte — funciona como um sinal antecedente de retencao de longo prazo (92% de D30 vs 22% na
ausencia de ambos), porque cria um gatilho diario externo ao app que reativa o usuario
independente do resultado do jogo.

**Acao:** Inserir um CTA de "jogue agora" na newsletter diaria, posicionado logo apos a secao de
manchetes, e medir via teste A/B (grupo controle sem CTA vs. grupo com CTA) o impacto no
`played_next_day` e no `active_d30` em uma coorte de 30 dias.

**Criterio de sucesso:** Saberei que funcionou quando o grupo exposto ao CTA apresentar D30
estatisticamente maior que o controle (teste qui-quadrado, p < 0.05) e um uplift absoluto de pelo
menos 5 p.p. em `played_next_day`, sustentado por 2 ciclos de 30 dias consecutivos.
""")

md("""## Anexo — arquivos e reproducao

| Script | Funcao |
|---|---|
| `01_explore.py` | Primeira leitura e shape/nulos/duplicatas |
| `02_quality.py` | Investigacao de dominio de valores |
| `03_quality2.py` | Datas mistas, cross-check attempts, confirmacao de mojibake na origem |
| `04_clean_and_analyze.py` | Limpeza documentada -> `*_clean.csv` + `cleaning_log.txt` |
| `05_stats_analysis.py` | Testes estatisticos completos -> `stats_results.txt` |
| `06_dashboard.py` | Dashboard Excel -> `dashboard_palavritas.xlsx` |
| `07_build_report.py` | Relatorio Word -> `Analise_Palavritas_the_news.docx` |
| `08_attempts_analysis.py` | Analise complementar da tabela attempts + efeitos compostos D+1 -> `attempts_analysis_results.txt` |

Para reproduzir do zero: rode os scripts na ordem numerica, ou execute este notebook celula a celula.
""")

nb['cells'] = cells
with open('Analise_Palavritas_the_news.ipynb', 'w', encoding='utf-8') as f:
    nbf.write(nb, f)
print("notebook written")
