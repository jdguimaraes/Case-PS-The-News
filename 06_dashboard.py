import pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

s = pd.read_csv('sessions_clean.csv')
u = pd.read_csv('user_profile_clean.csv')
s['played_next_day'] = s['played_next_day'].astype(bool)
s['active_d30'] = s['active_d30'].astype(bool)
s['newsletter_open_before_game'] = s['newsletter_open_before_game'].astype(bool)
s['hour_bucket'] = pd.cut(s['session_hour'], bins=[5,9,12,17,21,24],
    labels=['Manha cedo (6-9h)','Manha (10-12h)','Tarde (13-17h)','Noite (18-21h)','Madrugada (22-23h)'])

NAVY = '1F3A5F'
BLUE = '2E5C8A'
LIGHT = 'EAF1F8'
GREEN = '2E8B57'
RED = 'B33A3A'
WHITE = 'FFFFFF'

wb = Workbook()

def style_title(ws, cell, text, size=16, color=NAVY):
    c = ws[cell]; c.value = text
    c.font = Font(bold=True, size=size, color=color)

def style_subtitle(ws, cell, text, color='666666'):
    c = ws[cell]; c.value = text
    c.font = Font(italic=True, size=10, color=color)

def kpi_card(ws, top_left, label, value, sub='', fill=BLUE):
    col0, row0 = top_left
    for r in range(row0, row0+4):
        for col in range(col0, col0+3):
            cell = ws.cell(r, col)
            cell.fill = PatternFill('solid', fgColor=fill)
    lc = ws.cell(row0, col0, label)
    lc.font = Font(bold=True, size=11, color=WHITE)
    lc.alignment = Alignment(wrap_text=True, vertical='top')
    vc = ws.cell(row0+1, col0, value)
    vc.font = Font(bold=True, size=28, color=WHITE)
    vc.alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=row0+1, start_column=col0, end_row=row0+2, end_column=col0+2)
    sc = ws.cell(row0+3, col0, sub)
    sc.font = Font(size=9, italic=True, color=WHITE)
    sc.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row0+3, start_column=col0, end_row=row0+3, end_column=col0+2)

def write_table(ws, df, start_cell, title=None):
    col0_letter, row0 = start_cell[0], start_cell[1]
    col0 = ws[f"{col0_letter}{row0}"].column
    r = row0
    if title:
        ws.cell(r, col0, title).font = Font(bold=True, size=12, color=NAVY)
        r += 2
    hdr_row = r
    for j, colname in enumerate(df.columns):
        c = ws.cell(hdr_row, col0+j, colname)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill('solid', fgColor=BLUE)
        c.alignment = Alignment(horizontal='center')
    for i, row in enumerate(df.itertuples(index=False), start=1):
        for j, val in enumerate(row):
            cell = ws.cell(hdr_row+i, col0+j, val)
            cell.alignment = Alignment(horizontal='center' if j>0 else 'left')
            if i % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=LIGHT)
    return hdr_row, hdr_row+len(df)

def add_bar(ws, anchor, data_ref, cat_ref, title, color=BLUE):
    chart = BarChart()
    chart.title = title
    chart.style = 10
    chart.y_axis.title = '%'
    chart.height = 8
    chart.width = 16
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.dLabls = DataLabelList(); chart.dLabls.showVal = True
    s_ = chart.series[0]
    s_.graphicalProperties.solidFill = color
    ws.add_chart(chart, anchor)
    return chart

# ============ SHEET 1: VISAO GERAL ============
ws0 = wb.active
ws0.title = "Visao Geral"
ws0.sheet_view.showGridLines = False
style_title(ws0, 'B2', 'Palavritas - O que explica a retencao')
style_subtitle(ws0, 'B3', '39.849 sessoes validas | 1.200 usuarios | 800 com perfil de pesquisa')

kpi_card(ws0, (2,5), 'Retorno no dia seguinte (D+1)\nmedia geral', f"{s['played_next_day'].mean()*100:.1f}%",
         'Nenhuma variavel testada explica isso de forma relevante', fill=RED)
kpi_card(ws0, (6,5), 'Ativo em 30 dias (D30)\nmedia geral', f"{s['active_d30'].mean()*100:.1f}%",
         'Drivers claros: D+1, manha, newsletter', fill=BLUE)
combo = s.groupby(['newsletter_open_before_game','played_next_day'])['active_d30'].mean()
kpi_card(ws0, (10,5), 'D30 quando abre newsletter\nE volta no D+1', f"{combo[(True,True)]*100:.1f}%",
         'vs. 21.8% de quem nao faz nenhuma das duas', fill=GREEN)

for col in 'BCDEFGHIJKLM':
    ws0.column_dimensions[col].width = 11
ws0.row_dimensions[5].height = 50
ws0.row_dimensions[6].height = 20
ws0.row_dimensions[7].height = 20
ws0.row_dimensions[8].height = 26

t = pd.DataFrame({
    'O que testamos': ['Horario do jogo', 'Palavra do dia', 'Device', 'Setor', 'Salario',
                       'Resultado (ganhou/perdeu)', 'Newsletter aberta no dia', '-> Voltou no D+1 (efeito em cascata p/ D30)',
                       '-> Jogou de manha (p/ D30)', '-> Abriu newsletter (p/ D30)'],
    'Significativo?': ['Nao (D+1)','Nao (D+1)','Nao (D+1 e D30)','Nao (D+1 e D30)','Nao (D+1 e D30)',
                        'Pequeno (D30)','Nao p/D+1, SIM p/D30','SIM - maior efeito','SIM','SIM'],
})
write_table(ws0, t, ('B', 16), 'Resumo dos testes estatisticos (qui-quadrado / teste-t, 95% confianca)')

# ============ SHEET 2: RETENCAO D30 ============
ws1 = wb.create_sheet("Retencao D30 (drivers)")
ws1.sheet_view.showGridLines = False
style_title(ws1, 'B2', 'O que explica ficar ativo depois de 30 dias', size=14)
style_subtitle(ws1, 'B3', 'Os 3 fatores abaixo tem efeito grande e estatisticamente significativo (p < 0.001)')

t1 = s.groupby('hour_bucket')['active_d30'].agg(['mean','count']).reset_index()
t1.columns = ['Horario','Taxa Ativo D30 (%)','N sessoes']
t1['Taxa Ativo D30 (%)'] = (t1['Taxa Ativo D30 (%)']*100).round(1)
r0, r1 = write_table(ws1, t1, ('B', 6), '1) Retencao D30 por horario de jogo')
add_bar(ws1, 'F6', Reference(ws1, min_col=3, min_row=r0, max_row=r1), Reference(ws1, min_col=2, min_row=r0+1, max_row=r1),
        'Taxa Ativo D30 (%) por horario', color=BLUE)

t2 = s.groupby('newsletter_open_before_game')['active_d30'].agg(['mean','count']).reset_index()
t2.columns = ['Abriu newsletter antes?','Taxa Ativo D30 (%)','N sessoes']
t2['Abriu newsletter antes?'] = t2['Abriu newsletter antes?'].map({True:'Sim', False:'Nao'})
t2['Taxa Ativo D30 (%)'] = (t2['Taxa Ativo D30 (%)']*100).round(1)
r0, r1 = write_table(ws1, t2, ('B', 16), '2) Retencao D30 por abertura de newsletter antes do jogo')
add_bar(ws1, 'F16', Reference(ws1, min_col=3, min_row=r0, max_row=r1), Reference(ws1, min_col=2, min_row=r0+1, max_row=r1),
        'Taxa Ativo D30 (%) - abriu newsletter?', color=GREEN)

t3 = s.groupby('played_next_day')['active_d30'].agg(['mean','count']).reset_index()
t3.columns = ['Voltou no D+1?','Taxa Ativo D30 (%)','N sessoes']
t3['Voltou no D+1?'] = t3['Voltou no D+1?'].map({True:'Sim', False:'Nao'})
t3['Taxa Ativo D30 (%)'] = (t3['Taxa Ativo D30 (%)']*100).round(1)
r0, r1 = write_table(ws1, t3, ('B', 26), '3) Retencao D30 por retorno no D+1 (maior efeito isolado)')
add_bar(ws1, 'F26', Reference(ws1, min_col=3, min_row=r0, max_row=r1), Reference(ws1, min_col=2, min_row=r0+1, max_row=r1),
        'Taxa Ativo D30 (%) - voltou D+1?', color=RED)

combo_df = combo.reset_index()
combo_df.columns = ['Abriu newsletter','Voltou D+1','Taxa Ativo D30 (%)']
combo_df['Abriu newsletter'] = combo_df['Abriu newsletter'].map({True:'Sim', False:'Nao'})
combo_df['Voltou D+1'] = combo_df['Voltou D+1'].map({True:'Sim', False:'Nao'})
combo_df['Taxa Ativo D30 (%)'] = (combo_df['Taxa Ativo D30 (%)']*100).round(1)
r0, r1 = write_table(ws1, combo_df, ('B', 36), '4) Efeito combinado (o achado principal): newsletter + D+1 juntos')

for col, w in zip('BCDEFGHIJ', [22,18,12,4,18,18,18,18,18]):
    ws1.column_dimensions[col].width = w

# ============ SHEET 3: D+1 (sem efeito) ============
ws2 = wb.create_sheet("Retorno D+1 (sem driver)")
ws2.sheet_view.showGridLines = False
style_title(ws2, 'B2', 'O que NAO explica o retorno no dia seguinte', size=14)
style_subtitle(ws2, 'B3', 'Taxa geral fica ~22% em praticamente todos os cortes testados (p > 0.05)')

row_cursor = 6
for col, label in [('hour_bucket','Horario'), ('device','Device'), ('result','Resultado')]:
    tt = s.groupby(col)['played_next_day'].agg(['mean','count']).reset_index()
    tt.columns = [label, 'Taxa D+1 (%)', 'N sessoes']
    tt['Taxa D+1 (%)'] = (tt['Taxa D+1 (%)']*100).round(1)
    r0, r1 = write_table(ws2, tt, ('B', row_cursor), f'Retencao D+1 por {label}')
    add_bar(ws2, f'F{row_cursor}', Reference(ws2, min_col=3, min_row=r0, max_row=r1), Reference(ws2, min_col=2, min_row=r0+1, max_row=r1),
            f'Taxa D+1 (%) por {label}', color='888888')
    row_cursor = r1 + 12

for col, w in zip('BCDEFGHIJ', [22,18,12,4,18,18,18,18,18]):
    ws2.column_dimensions[col].width = w

# ============ SHEET 4: DADOS (apoio) ============
ws3 = wb.create_sheet("Dados de apoio")
ws3.sheet_view.showGridLines = False
style_title(ws3, 'B2', 'Tabelas de apoio - perfil do usuario', size=14)
su = s.merge(u, on='user_id', how='left')
row_cursor = 6
for col, label, target, tname in [
    ('sector','Setor','active_d30','D30'), ('salary_range','Faixa salarial','active_d30','D30'),
    ('orders_food_delivery','Pede delivery?','active_d30','D30')]:
    tt = su.dropna(subset=[col]).groupby(col)[target].agg(['mean','count']).reset_index()
    tt.columns = [label, f'Taxa {tname} (%)', 'N sessoes']
    tt[f'Taxa {tname} (%)'] = (tt[f'Taxa {tname} (%)']*100).round(1)
    r0, r1 = write_table(ws3, tt, ('B', row_cursor), f'{tname} por {label} (sem efeito significativo)')
    row_cursor = r1 + 3

for col, w in zip('BCDEFGHIJ', [22,18,12]):
    ws3.column_dimensions[col].width = w

wb.save('dashboard_palavritas.xlsx')
print("saved improved dashboard")
