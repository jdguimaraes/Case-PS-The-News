import pandas as pd, re
path = r"../Case Analista de Dados — the news [Dataset Palavritas].xlsx"
xls = pd.read_excel(path, sheet_name=None)
s = xls['palavritas_sessions']
a = xls['palavritas_attempts']
u = xls['user_profile']

print("### word_date format patterns ###")
def fmt(x):
    if re.match(r'^\d{4}-\d{2}-\d{2}$', x): return 'YYYY-MM-DD'
    if re.match(r'^\d{2}/\d{2}/\d{4}$', x): return 'DD/MM/YYYY or MM/DD/YYYY'
    return 'other:'+x
print(s['word_date'].apply(fmt).value_counts())
print(s[s['word_date'].apply(fmt)!='YYYY-MM-DD']['word_date'].head(10).tolist())

print("\n### attempt count per session vs sessions.attempts field ###")
real_counts = a.groupby('session_id')['attempt_number'].max().rename('real_attempts')
merged = s[['session_id','attempts','result']].merge(real_counts, on='session_id', how='left')
merged['real_attempts'] = merged['real_attempts'].fillna(0)
mismatch = merged[merged['attempts'] != merged['real_attempts']]
print("mismatched rows:", len(mismatch), "of", len(merged))
print(mismatch.head(10))

print("\n### attempts field distribution vs valid 1-6 ###")
print(s['attempts'].value_counts().sort_index())

print("\n### checking mojibake - raw bytes via openpyxl ###")
from openpyxl import load_workbook
wb = load_workbook(path, read_only=True)
ws = wb['user_profile']
rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
for r in rows:
    print(r)

print("\n### encoding check: try latin1->utf8 fix on sector ###")
def fix_enc(x):
    if not isinstance(x,str): return x
    try:
        return x.encode('latin1').decode('utf-8')
    except Exception:
        return x
print(u['sector'].dropna().unique())
print([fix_enc(v) for v in u['sector'].dropna().unique()])
print([fix_enc(v) for v in u['state'].dropna().unique()])

print("\n### duplicate full rows detail (same session_id) - are they byte-identical? ###")
dup = s[s.duplicated(subset=['session_id'], keep=False)].sort_values('session_id')
def all_identical(d):
    return d.drop(columns=['session_id']).duplicated(keep=False).all()
g = dup.groupby('session_id', group_keys=False).apply(all_identical, include_groups=False)
print("all-identical duplicate groups:", g.sum(), "of", g.shape[0])
print(g[~g].head(5))
